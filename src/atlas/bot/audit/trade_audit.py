# src/atlas/bot/audit/trade_audit.py
from __future__ import annotations

from typing import List, Dict, Any, Optional
import os

from openpyxl import Workbook, load_workbook

from atlas.bot.atlas_ia.state import AtlasIAState, TradeVirtual


def _now_ts() -> int:
    import time
    return int(time.time())


def _r_multiple(side: str, entry: float, sl: float, exit_price: float) -> float:
    """
    R = (ganancia en precio) / (riesgo en precio)
    BUY: riesgo = entry - sl
    SELL: riesgo = sl - entry
    """
    if side == "BUY":
        risk = entry - sl
        if risk == 0:
            return 0.0
        return (exit_price - entry) / risk
    else:
        risk = sl - entry
        if risk == 0:
            return 0.0
        return (entry - exit_price) / risk


def _risk_usd(balance: float, risk_pct: float) -> float:
    return balance * (risk_pct / 100.0)


def ensure_excel(path: str) -> None:
    if os.path.exists(path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "trades"

    ws.append([
        "#",
        "Par",
        "Fecha",
        "Hora",
        "Mundo",
        "Dirección",
        "Entrada",
        "SL",
        "TP",
        "Riesgo%",
        "BalanceInicio",
        "Salida",
        "MotivoSalida",
        "R",
        "$",
        "Feedback",
    ])
    wb.save(path)


def append_trade_to_excel(path: str, t: TradeVirtual) -> None:
    ensure_excel(path)
    wb = load_workbook(path)
    ws = wb["trades"]

    import datetime
    dt_open = datetime.datetime.fromtimestamp(t.open_ts)
    fecha = dt_open.strftime("%Y-%m-%d")
    hora = dt_open.strftime("%H:%M:%S")

    ws.append([
        t.trade_id,
        t.symbol,
        fecha,
        hora,
        f"ATLAS_IA:{t.atlas_mode}",
        t.side,
        round(t.entry, 5),
        round(t.sl, 5),
        round(t.tp, 5),
        t.risk_pct,
        round(t.balance_start, 2),
        None if t.exit_price is None else round(t.exit_price, 5),
        t.exit_reason,
        None if t.result_r is None else round(t.result_r, 3),
        None if t.result_usd is None else round(t.result_usd, 2),
        t.feedback,
    ])

    wb.save(path)


def simulate_trade_on_candles(
    state: AtlasIAState,
    trade: TradeVirtual,
    candles: List[Dict[str, Any]],
    sl_first: bool = True,
) -> Optional[TradeVirtual]:
    """
    Revisa candles nuevas y cierra trade si toca SL o TP.
    Regla conservadora: si una vela toca ambos y sl_first=True => SL.
    """
    if trade.status != "OPEN":
        return trade

    for c in candles:
        high = float(c["high"])
        low = float(c["low"])
        close = float(c["close"])
        ts = int(c["time"])

        if trade.side == "BUY":
            hit_sl = low <= trade.sl
            hit_tp = high >= trade.tp
        else:
            hit_sl = high >= trade.sl
            hit_tp = low <= trade.tp

        if hit_sl and hit_tp:
            if sl_first:
                return close_trade(state, trade, trade.sl, "SL", ts)
            else:
                return close_trade(state, trade, trade.tp, "TP", ts)

        if hit_sl:
            return close_trade(state, trade, trade.sl, "SL", ts)

        if hit_tp:
            return close_trade(state, trade, trade.tp, "TP", ts)

    return trade


def close_trade(state: AtlasIAState, trade: TradeVirtual, exit_price: float, reason: str, close_ts: Optional[int] = None) -> TradeVirtual:
    if close_ts is None:
        close_ts = _now_ts()

    trade.exit_price = float(exit_price)
    trade.exit_reason = reason
    trade.close_ts = int(close_ts)
    trade.status = "CLOSED"

    r = _r_multiple(trade.side, trade.entry, trade.sl, trade.exit_price)
    trade.result_r = float(r)

    risk_usd = _risk_usd(trade.balance_start, trade.risk_pct)
    trade.result_usd = float(r * risk_usd)

    # Actualiza balance audit (equity “virtual”)
    state.audit_balance = float(state.audit_balance + trade.result_usd)

    # Export a Excel automático (solo cuando cierra)
    append_trade_to_excel(state.excel_path, trade)

    return trade