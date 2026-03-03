from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, Optional, Tuple
import os
import time

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_PATH = os.path.join(os.getcwd(), "atlas_journal.xlsx")
SHEET_NAME = "journal"


@dataclass
class JournalRow:
    ts: int
    symbol: str
    tf: str
    world: str
    phase: str
    action: str
    reason: str
    side: str
    entry: float
    sl: float
    tp: float
    doctrine_ok: bool
    scenario_closed: bool
    gap_state: str


_HEADERS = [
    "ts",
    "symbol",
    "tf",
    "world",
    "phase",
    "action",
    "reason",
    "side",
    "entry",
    "sl",
    "tp",
    "doctrine_ok",
    "scenario_closed",
    "gap_state",
]

# Anti-duplicado simple (por proceso)
_LAST_KEY: Optional[int] = None


def _ensure_book(path: str) -> Tuple[Workbook, Worksheet]:
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # Header
    if ws.max_row < 1 or (ws.cell(row=1, column=1).value != _HEADERS[0]):
        ws.delete_rows(1, ws.max_row)
        ws.append(_HEADERS)
    return wb, ws


def _hash_key(
    ts: int,
    symbol: str,
    tf: str,
    world: str,
    action: str,
    reason: str,
    side: str,
    entry: float,
    sl: float,
    tp: float,
) -> int:
    # Redondeos para evitar “ruido” de floats por tick
    e = round(float(entry), 6)
    s = round(float(sl), 6)
    t = round(float(tp), 6)
    base = f"{ts}|{symbol}|{tf}|{world}|{action}|{reason}|{side}|{e}|{s}|{t}"
    return hash(base)


def maybe_log_trade(
    snap: Dict[str, Any],
    *,
    path: str = DEFAULT_PATH,
) -> Dict[str, Any]:
    """
    Registra SOLO si:
    - decision.action == 'TRADE'
    - doctrine_ok == True
    - anti-dup OK
    NO ejecuta nada. Solo journal.
    """
    global _LAST_KEY

    decision = (snap or {}).get("decision") or {}
    bot_state = (snap or {}).get("bot_state") or {}
    gap_info = (snap or {}).get("gap_info") or {}
    scenario = (bot_state or {}).get("scenario") or (snap or {}).get("scenario") or {}

    action = str(decision.get("action", "NO_TRADE")).upper()
    doctrine_ok = bool(snap.get("doctrine_ok", True))

    if action != "TRADE":
        return {"logged": False, "why": "NOT_TRADE"}

    if not doctrine_ok:
        return {"logged": False, "why": "DOCTRINE_FAIL"}

    ts = int(snap.get("ts", 0) or time.time())
    symbol = str(snap.get("symbol", ""))
    tf = str(snap.get("tf", ""))
    world = str(snap.get("world", "GENERAL"))
    phase = str(bot_state.get("phase", "IDLE"))
    reason = str(decision.get("reason", ""))[:120]

    side = str(decision.get("side", ""))[:8]
    entry = float(decision.get("entry", 0.0) or 0.0)
    sl = float(decision.get("sl", 0.0) or 0.0)
    tp = float(decision.get("tp", 0.0) or 0.0)

    scenario_closed = bool(scenario.get("closed", False))
    gap_state = str((gap_info or {}).get("state", "") or (bot_state.get("gap", {}) or {}).get("state", ""))

    key = _hash_key(ts, symbol, tf, world, action, reason, side, entry, sl, tp)
    if _LAST_KEY == key:
        return {"logged": False, "why": "DUPLICATE"}

    wb, ws = _ensure_book(path)
    ws.append([
        ts,
        symbol,
        tf,
        world,
        phase,
        action,
        reason,
        side,
        entry,
        sl,
        tp,
        doctrine_ok,
        scenario_closed,
        gap_state,
    ])
    wb.save(path)

    _LAST_KEY = key
    return {"logged": True, "path": path}


def journal_stats(path: str = DEFAULT_PATH) -> Dict[str, Any]:
    if not os.path.exists(path):
        return {"ok": True, "path": path, "rows": 0, "last_ts": 0}

    wb = load_workbook(path, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        return {"ok": True, "path": path, "rows": 0, "last_ts": 0}

    ws = wb[SHEET_NAME]
    rows = max(0, ws.max_row - 1)
    last_ts = 0
    if ws.max_row >= 2:
        last_ts = int(ws.cell(row=ws.max_row, column=1).value or 0)
    return {"ok": True, "path": path, "rows": rows, "last_ts": last_ts}